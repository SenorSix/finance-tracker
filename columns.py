import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

# Load your cleaned data
file_path = "excel/cleaned_descriptions.csv"
df = pd.read_csv(file_path)

# Convert date column
df["Date"] = pd.to_datetime(df["Date"])

# Clean dollar signs and convert to float
df[["Amount", "Balance"]] = df[["Amount", "Balance"]].replace('[\$,]', '', regex=True).astype(float)

# This pivot table is for the totals row. pivot_df below is for all the transaction rows enabling multiple vendor transactions in a single day
totals_pivot_df = df.pivot_table(
    index="Date",
    columns="Cleaned_Description_1",
    values="Amount",
    aggfunc="sum",
    fill_value=0
)

# Create pivot table (Date as index, Vendors as columns, Amounts as values)
pivot_df = df.pivot_table(
    index="Date",
    columns="Cleaned_Description_1",
    values="Amount",
    aggfunc=lambda x: x.iloc[0] if len(x) == 1 else " + ".join(f"{v:.2f}" for v in x),  # Doesn't sum in order to show multiple purchases from vendor in one day
    fill_value=""
)

# Build full date range from earliest to latest date
full_range = pd.date_range(start=df["Date"].min(), end=df["Date"].max(), freq='D')

# Reindex to include all days (even if no transactions)
pivot_df = pivot_df.reindex(full_range)

# Reset index so "Date" becomes a column instead of index
pivot_df = pivot_df.reset_index()
pivot_df.columns.name = None  # Remove the 'Cleaned_Description_1' header

# Format date
pivot_df["index"] = pivot_df["index"].dt.strftime("%-m-%-d-%Y")  # Use %#m/%#d on Windows if needed
pivot_df.rename(columns={"index": "Date"}, inplace=True)

# Add totals row at the bottom
totals = totals_pivot_df.sum()
# TODO: Pandas may change downcasting rules soon — revisit this if it breaks

totals_row = ["Total"] + totals.tolist()
pivot_df.loc[len(pivot_df)] = totals_row

# Save to Excel
excel_path = "excel/output_by_date.xlsx"
pivot_df.to_excel(excel_path, index=False)

# Auto-adjust column widths
wb = load_workbook(excel_path)
ws = wb.active
ws.freeze_panes = "B2"
last_row = ws.max_row

scaling_factor = 1.1
padding = 2
max_column_width = 40

for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                length = len(str(cell.value).strip())
                if length < 100:
                    max_length = max(max_length, length)
        except:
            pass
    adjusted_width = min((max_length * scaling_factor) + padding, max_column_width)
    ws.column_dimensions[col_letter].width = adjusted_width

# Bold and border the totals row
bold_border = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

for cell in ws[last_row]:
    cell.font = Font(bold=True)
    cell.border = bold_border

wb.save(excel_path)

print("✅ Output with full date range and totals row saved!")
